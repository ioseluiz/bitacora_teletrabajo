import os
import sys
import win32com.client
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class ReportExporter:
    def __init__(self, template_path='templates/template.xlsx', base_dir='data/reports'):
        self.template_path = resource_path(template_path)
        self.set_output_dir(base_dir)

    def set_output_dir(self, base_dir):
        self.output_dir_xlsx = os.path.join(base_dir, 'xlsx')
        self.output_dir_pdf = os.path.join(base_dir, 'pdf')
        os.makedirs(self.output_dir_xlsx, exist_ok=True)
        os.makedirs(self.output_dir_pdf, exist_ok=True)

    def _write_to_cell(self, ws, cell_coords, value, center=False):
        """Writes value to a cell, even if it is part of a merged range."""
        from openpyxl.cell.cell import MergedCell
        cell = ws[cell_coords]
        target_cell = cell
        
        if isinstance(cell, MergedCell):
            # Find the master cell of the merged range
            for merged_range in ws.merged_cells.ranges:
                if cell_coords in merged_range:
                    target_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        
        target_cell.value = value
        if center:
            target_cell.alignment = Alignment(horizontal='center', vertical='center')

    def export(self, report_data, tasks):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_base = f"Reporte_{report_data['date']}_{timestamp}"
        xlsx_path = os.path.abspath(os.path.join(self.output_dir_xlsx, f"{filename_base}.xlsx"))
        pdf_path = os.path.abspath(os.path.join(self.output_dir_pdf, f"{filename_base}.pdf"))

        wb = load_workbook(self.template_path)
        ws = wb.active

        # Rename sheet to YYYYMMDD based on report date (dd-mm-yyyy)
        try:
            date_obj = datetime.strptime(report_data['date'], "%d-%m-%Y")
            ws.title = date_obj.strftime("%Y%m%d")
        except:
            pass 

        # Mapping as requested using helper
        self._write_to_cell(ws, 'B5', report_data['date'])
        self._write_to_cell(ws, 'D5', report_data['pay_period'])
        self._write_to_cell(ws, 'B6', report_data['employee_name'])
        self._write_to_cell(ws, 'D6', report_data['ip'])
        self._write_to_cell(ws, 'B7', report_data['office'])
        self._write_to_cell(ws, 'D7', report_data.get('job_title', ""))
        self._write_to_cell(ws, 'B8', report_data['supervisor_name'])
        self._write_to_cell(ws, 'D8', report_data.get('supervisor_title', ""))
        self._write_to_cell(ws, 'B9', report_data['schedule'])
        self._write_to_cell(ws, 'D9', report_data['lunch_period'])

        # Cleanup rows 19 to 23 (Columns A to D)
        for row in range(19, 24):
            for col in ['A', 'B', 'C', 'D']:
                self._write_to_cell(ws, f'{col}{row}', None)

        # Tasks - Rows 12 to 17
        # Description: A (Merged with B), Resources: C, Duration: D (Centered)
        for i, task in enumerate(tasks):
            if i > 5: break 
            row = 12 + i
            self._write_to_cell(ws, f'A{row}', task['description'])
            self._write_to_cell(ws, f'C{row}', task['resources'])
            self._write_to_cell(ws, f'D{row}', task['duration'], center=True)
        
        # Total - D24
        self._write_to_cell(ws, 'D24', report_data['total_hours'], center=True)

        wb.save(xlsx_path)
        self._convert_to_pdf(xlsx_path, pdf_path)

        return xlsx_path, pdf_path

    def _convert_to_pdf(self, xlsx_path, pdf_path):
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        try:
            wb = excel.Workbooks.Open(xlsx_path)
            # xlTypePDF = 0
            wb.ExportAsFixedFormat(0, pdf_path)
            wb.Close(False)
        finally:
            excel.Quit()
